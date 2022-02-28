################################# Equal power allocation with mobile users ####################################
import pandas as pd
from openpyxl import workbook,load_workbook
from  random import randrange,uniform
from  random import randrange,uniform
import numpy as np
from math import cos,sin,pi,log10,log2
import pylab
################### Initializing position of rx and tx in circular form #########################
def loc_init(d2d_dist, num_d2d):
    r=100         # radius of circle
    rx_loc = np.zeros((num_d2d , 2))
    tx_loc = np.zeros((num_d2d , 2))
    for i in range(num_d2d):
        angle= randrange(0,360)
        angle=angle*pi/180
        tx_loc[i, 0] = r*cos(angle)
        rx_loc[i, 0] = r*cos(angle) + d2d_dist
        tx_loc[i, 1] = r*sin(angle)
        rx_loc[i,1] = r*sin(angle)    
    return rx_loc, tx_loc
def get_loc(rx_loc, tx_loc,num_d2d) :
  for i in range(num_d2d):
        angle= randrange(0,360)
        r = uniform(0.8,2.8)
        angle=angle*pi/180
        tx_loc[i, 0] += r*cos(angle)
        tx_loc[i, 1] += r*sin(angle)
  for i in range(num_d2d):
        angle= randrange(0,360)
        r = uniform(0.8,2.8)
        angle=angle*pi/180
        rx_loc[i, 0] += r*cos(angle)
        rx_loc[i, 1] += r*sin(angle)
  return rx_loc, tx_loc
def ch_gen(rx_loc, tx_loc, num_d2d,C_frequency):
    
    ch_w_fading = np.zeros((num_d2d,num_d2d))
    multi_fading = 0.5 * np.random.randn(num_d2d, num_d2d) ** 2 + 0.5 * np.random.randn(num_d2d, num_d2d) ** 2
    for i in range(num_d2d) :         
      
      for j in range(num_d2d) :
          
          d =  np.linalg.norm(tx_loc[i] - rx_loc[j])                                        # distance vector has value of distance between each tx and rx                         
          if d<18:
            pu_ch_gain_db= 28 + 22 * log10(d) + 20 * log10(C_frequency)       # LOS pathloss
            pu_ch_gain = 10**(-pu_ch_gain_db/10)
          
          else :
            pu_ch_gain_db = 32.4 + 30 * log10(d) + 20 * log10(C_frequency)    # NLOS pathloss 
            pu_ch_gain = 10**(-pu_ch_gain_db/10)
          multi_fading = 0.5 * np.random.randn(num_d2d, num_d2d) ** 2 + 0.5 * np.random.randn(num_d2d, num_d2d) ** 2
          final_ch = (pu_ch_gain *(np.sum(multi_fading) ))
         # print(final_ch )
          ch_w_fading[i,j] = final_ch    
    return  ch_w_fading 
####################### simulation parameters ########################
bw = 1*10**9              # Bandwidth = 1GHz
p_t_dB = 23               # maximum d2d transmit power in dB
p_t = 10**(p_t_dB/10)      
C_frequency = 28e9         # carrier frequency = 28 GHz
num_d2d =4
Noise = -174               # noise = -174 dBm
s= 10**(-174/10)
d2d_dist =50

###############################  For creating excel sheets of value ####################################
wb1=load_workbook('receiver1.xlsx')          
ws1 = wb1.active
wb2=load_workbook('receiver2.xlsx')
ws2 = wb2.active
wb3=load_workbook('receiver3.xlsx')
ws3 = wb3.active
wb4=load_workbook('receiver4.xlsx')
ws4 = wb4.active
              

for iteration in range(100) :
  rx_loc, tx_loc = loc_init(d2d_dist, num_d2d)           # initializing tx-rx pairs location 
  #########################  Assigning equal power ###############
  d2d_power =np.ones(num_d2d)*(23/4)
  sinr = np.zeros(num_d2d)
  cap = np.zeros(num_d2d)

  for sim_time  in range(1000) :
    rx_loc, tx_loc = get_loc(rx_loc, tx_loc,num_d2d)
    for i in range(num_d2d) :
      for j in range(num_d2d) :
        d =  np.linalg.norm(tx_loc[i] - rx_loc[j])  
        if d < 100 :
          if j==0 :
           x = h[i,j]*d2d_power[i]/s
           sinr[j] = 10*log10(x)
           cap[j] += bw*log2(1+x)

          elif j==1 :
            x = h[i,j]*d2d_power[i]/s
            sinr[j] = 10*log10(x)
            cap[j] += bw*log2(1+x)

          elif j==2 :
            x = h[i,j]*d2d_power[i]/s
            sinr[j] = 10*log10(x)
            cap[j] += bw*log2(1+x)

          elif j==3 :
            x = h[i,j]*d2d_power[i]/s
            sinr[j] = 10*log10(x)
            cap[j] += bw*log2(1+x)

  for k in range(num_d2d) :
    if k==0 :
      ws1.append([cap[k]])
    elif k==1:
       ws2.append([cap[k]])
    elif k==2:
       ws3.append([cap[k]])
    elif k==3:
      ws4.append([cap[k]])

wb1.save('receiver1.xlsx')
wb2.save('receiver2.xlsx')
wb3.save('receiver3.xlsx')
wb4.save('receiver4.xlsx')