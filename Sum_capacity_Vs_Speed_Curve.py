############################## For first curve ################################
import pandas as pd
from openpyxl import workbook,load_workbook
import numpy as np
from math import log10,log2,sqrt,pi
from mpmath import besselj
from random import randrange,random
from matplotlib import pyplot as plt
################################## Simulation Parameters ###########################
f_c = 28                       # centre frequency is 28 GHz
AntGain = 3                # vehicle antenna gain 3 dBi
NoiseFigure = 9            # vehicle noise figure 9 dB
dB_noise = -114;           # noise power in dBm
P_max_dB = 23                 # d2d maximum power
T = 1
num_d2d = 2


################### dB to linear scale conversion ###############
noise = 10**(dB_noise/10)
P_max =  10**(P_max_dB/10)

################### Algorithm defination #################
def WMMSE(num_d2d, H, Pmax, var_noise):
    k = num_d2d
    vnew = 0
    b = np.ones((k))*(sqrt(Pmax))
    f = np.zeros((k))
    w = np.zeros((k))
    for i in range(k):
        f[i] = H[i,i] * b[i] /(np.square(H[i,:]) @ np.square(b) + var_noise)
        w[i] = 1 / (1 - f[i] * b[i] * H[i, i])
        vnew = vnew + log2(w[i])

    VV = np.zeros(100)
    for iter in range(100):
        vold = vnew
        for i in range(k):
            btmp = w[i] * f[i] * H[i, i] / sum(w * np.square(f) * np.square(H[:, i]))
            b[i] = min(btmp, np.sqrt(Pmax)) + max(btmp, 0) - btmp

        vnew = 0
        for i in range(k):
            f[i] = H[i, i] * b[i] / ((np.square(H[i, :])) @ (np.square(b)) + var_noise)
            w[i] = 1 / (1 - f[i] * b[i] * H[i, i])
            vnew = vnew + log2(w[i])

        VV[iter] = vnew
        if vnew - vold <= 1e-3:
            break

    p_opt = np.square(b)
    return p_opt

######################## algorithm end ####################################

################################## Path Loss #########################################
def pathloss(f_c,d) :                                                             
    if d<18:
       pathloss= 28 + 22 * log10(d) + 20 * log10(f_c)       # LOS pathloss
       PL_dB = -(np.random.randn()*3 + pathloss) + 2*AntGain - NoiseFigure 
    else :
       pathloss = 32.4 + 30 * log10(d) + 20 * log10(f_c)    # NLOS pathloss 
       PL_dB = -(np.random.randn()*4 + pathloss) + 2*AntGain - NoiseFigure
    return PL_dB

################### Initializing position of rx and tx in circular form #########################
def loc_init(num_d2d):
    r=100         # radius of circle
    rx_loc = np.zeros((num_d2d , 2))
    tx_loc = np.zeros((num_d2d , 2))
    for i in range(num_d2d):
        tx_loc[i, 0] = randrange(-200,0)
        rx_loc[i, 0] = randrange(0,200)
        tx_loc[i, 1] = 43
        rx_loc[i,1] =  39    
    return rx_loc, tx_loc

########################################## tx and rx location when they are in different lane   ###########################
def get_loc(rx_loc, tx_loc,num_d2d,speed) :
  for i in range(num_d2d):
        r = speed*5/18
        tx_loc[i, 0] += r
  for i in range(num_d2d):
        r = speed*5/18
        rx_loc[i, 0] -= r
  return rx_loc, tx_loc
###############################  For creating excel sheets of value ####################################
wb1=load_workbook('receiver1.xlsx')          
ws1 = wb1.active
######################### main simulation #####################

for speed in range(3,61,3) :
 sum_capacity =0  
 for iteration in range(100) :
  rx_loc, tx_loc = loc_init(num_d2d)           # initializing tx-rx pairs location 
  cap = 0
  a=0
  epsi = besselj(0,2*pi*(T*1e-3)*(f_c*1e9)*(speed/3.6)/(3e8))
  for sim_time  in range(300) :
     rx_loc, tx_loc = get_loc(rx_loc, tx_loc,num_d2d,speed)
     h = (np.random.randn(2,2)+1j*np.random.randn(2,2))/sqrt(2)
     for i in range(num_d2d) :
        d =  np.linalg.norm(tx_loc[i] - rx_loc[i])
        if d < 100 :
          pl = pathloss(f_c,d)
          pl = 10**(pl/10)
          h[i,i] = (random()+1j*random())/sqrt(2) 
          e = sqrt(1-epsi**2)*(random()+1j*random())/sqrt(2)
          g_d = pl*(abs(epsi*h[i,i])**2 + abs(e)**2)
          SNR = P_max*g_d/noise
          cap +=log2(1+SNR)
          #a+=1

  sum_capacity += cap /300          
 sum_capacity =sum_capacity/100
 ws1.append([sum_capacity])
wb1.save('receiver1.xlsx')



######################################## For second Curve #################################
############# Pathloss and simulation parameter  are same as above ######################
################### Initializing position of rx and tx in circular form #########################
def loc_init(num_d2d):
    r=100         # radius of circle
    rx_loc = np.zeros((num_d2d , 2))
    tx_loc = np.zeros((num_d2d , 2))
    for i in range(num_d2d):
      if i ==0 :
        tx_loc[i, 0] = randrange(-200,0)
        rx_loc[i, 0] = tx_loc[i, 0] + 50
        tx_loc[i, 1] = 39
        rx_loc[i,1] =  39 
      else :
        tx_loc[i, 0] = randrange(0,200)
        rx_loc[i, 0] = tx_loc[i, 0] + 50
        tx_loc[i, 1] = 43
        rx_loc[i,1] =  43

    return rx_loc, tx_loc

########################################## tx and rx location when they are in different lane   ###########################
def get_loc(rx_loc, tx_loc,num_d2d,speed) :
  for i in range(num_d2d):
    if i == 0:
        r = speed*5/18
        tx_loc[i, 0] -= r
    else :
        r = speed*5/18
        rx_loc[i, 0] += r
  return rx_loc, tx_loc
###############################  For creating excel sheets of value ####################################
wb1=load_workbook('receiver1.xlsx')          
ws1 = wb1.active
######################### main simulation #####################

for speed in range(3,61,3) :
 sum_capacity =0  
 for iteration in range(100) :
  rx_loc, tx_loc = loc_init(num_d2d)           # initializing tx-rx pairs location 
  cap = 0
  a=0
  epsi_k = besselj(0,2*pi*(T*1e-3)*(f_c*1e9)*(speed/3.6)/(3e8))
  for sim_time  in range(1000) :
     rx_loc, tx_loc = get_loc(rx_loc, tx_loc,num_d2d,speed)
     h = (np.random.randn(2,2)+1j*np.random.randn(2,2))/sqrt(2)
     for i in range(num_d2d) :
        d =  np.linalg.norm(tx_loc[i] - rx_loc[i])
        if d < 100 :
          pl = pathloss(f_c,d)
          pl = 10**(pl/10)
          h[i,i] = (0.5 * np.random.randn() ** 2 + 0.5 * np.random.randn() ** 2) 
          e_k = sqrt(1-epsi_k**2)*(random()+1j*random())/sqrt(2)
          g_k_d = pl*(abs(epsi_k*h[i,i])**2 + abs(e_k)**2)
          SNR = P_max*g_k_d/noise
          cap +=log2(1+SNR)
          a+=1

  sum_capacity += cap/1000        
 sum_capacity =sum_capacity/100
 ws1.append([sum_capacity])
wb1.save('receiver1.xlsx')