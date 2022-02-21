from math import sqrt,log2
import pandas as pd
from openpyxl import workbook,load_workbook
###########################  WMMSE Alogarithm ##################################

bw = 1*10**9              # Bandwidth = 1GHz
p_t_dB = 23               # maximum d2d transmit power in dB
p_t = 10**(p_t_dB/10)      
C_frequency = 28e9         # carrier frequency = 28 GHz
num_d2d =4
Noise = -174               # noise = -174 dBm
s= 10**(-174/10)

################### Algorithm defination #################
def WMMSE(p_int, H, Pmax, var_noise):
    K = np.size(p_int)
    vnew = 0
    b = np.sqrt(p_int)
    f = np.zeros(K)
    w = np.zeros(K)
    for i in range(K):
        f[i] = H[i, i] * b[i] /(np.square(H[i, :]) @ np.square(b) + var_noise)
        w[i] = 1 / (1 - f[i] * b[i] * H[i, i])
        vnew = vnew + log2(w[i])

    VV = np.zeros(100)
    for iter in range(100):
        vold = vnew
        for i in range(K):
            btmp = w[i] * f[i] * H[i, i] / sum(w * np.square(f) * np.square(H[:, i]))
            b[i] = min(btmp, np.sqrt(Pmax)) + max(btmp, 0) - btmp

        vnew = 0
        for i in range(K):
            f[i] = H[i, i] * b[i] / ((np.square(H[i, :])) @ (np.square(b)) + var_noise)
            w[i] = 1 / (1 - f[i] * b[i] * H[i, i])
            vnew = vnew + log2(w[i])

        VV[iter] = vnew
        if vnew - vold <= 1e-3:
            break

    p_opt = np.square(b)
    return p_opt

######################## algorithm end ####################################

###############################  For creating excel sheets of value ####################################
wb1=load_workbook('receiver1.xlsx')
ws1 = wb1.active
wb2=load_workbook('receiver2.xlsx')
ws2 = wb2.active
wb3=load_workbook('receiver3.xlsx')
ws3 = wb3.active
wb4=load_workbook('receiver4.xlsx')
ws4 = wb4.active

for z in range (51) :                           # 51 iterations
  ##########################  channel matrix ########################
  h=np.zeros((num_d2d,num_d2d))
  p_int =np.ones(num_d2d)*((p_t)) 
  for i in range(num_d2d) :
    for j in range(num_d2d) :
      x = 0.5 * np.random.randn(num_d2d, num_d2d) ** 2 + 0.5 * np.random.randn(num_d2d, num_d2d) ** 2
      h[i,j]=np.sum(x)

  ####################### calculation of optimal power and capacity  ###################
  print(h)    
  p_wmmse =  WMMSE(p_int,h,p_t,1)
  print(p_wmmse)
  cap = np.zeros(num_d2d)
  for i in range(num_d2d):
     s = 1
     for j in range(num_d2d):
        if j!=i:
          s = s+h[i,j]**2*p_wmmse[j]
     cap[i]=log2(1+h[i,i]**2*p_wmmse[i]/s)
     if i==0 :
        ws1.append([cap[i]])
     elif i==1:
        ws2.append([cap[i]])
     elif i==2:
        ws3.append([cap[i]])
     elif i==3:
        ws4.append([cap[i]])

wb1.save('receiver1.xlsx')
wb2.save('receiver2.xlsx')
wb3.save('receiver3.xlsx')
wb4.save('receiver4.xlsx')