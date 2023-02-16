############################ WMMSE Algorithm with mobile users #######################################
import pandas as pd
from openpyxl import workbook,load_workbook
from  random import randrange,uniform
from  random import randrange,uniform
import numpy as np
from math import cos,sin,pi,log10,log2,sqrt
import pylab
from WMMSE_algorithm import WMMSE
######################## Simulation Parameters #################
bw = 1*10**9              # Bandwidth = 1GHz
p_t_dB = 23               # maximum d2d transmit power in dB
p_t = 10**(p_t_dB/10)      
C_frequency = 28e9         # carrier frequency = 28 GHz
num_d2d =4
Noise = -174               # noise = -174 dBm
s= 10**(-174/10)
d2d_dist = 50
################### Initializing position of rx and tx in circular form #########################
def loc_init(d2d_dist, num_d2d):
    r=100         # radius of circle
    rx_loc = np.zeros((num_d2d , 2))
    tx_loc = np.zeros((num_d2d , 2))
    for i in range(num_d2d):
        angle= randrange(0,360)
        angle=angle*pi/180
        tx_loc[i, 0] = r*cos(angle)
        rx_loc[i, 0] = r*cos(angle) + d2d_dist
        tx_loc[i, 1] = r*sin(angle)
        rx_loc[i,1] = r*sin(angle)    
    return rx_loc, tx_loc
def get_loc(rx_loc, tx_loc,num_d2d) :
  for i in range(num_d2d):
        angle= randrange(0,360)
        r = uniform(0.8,2.8)
        angle=angle*pi/180
        tx_loc[i, 0] += r*cos(angle)
        tx_loc[i, 1] += r*sin(angle)
  for i in range(num_d2d):
        angle= randrange(0,360)
        r = uniform(0.8,2.8)
        angle=angle*pi/180
        rx_loc[i, 0] += r*cos(angle)
        rx_loc[i, 1] += r*sin(angle)
  return rx_loc, tx_loc
################### Algorithm defination #################
def WMMSE(num_d2d, H, Pmax, var_noise):
    k = num_d2d
    vnew = 0
    b = np.ones((k))*(sqrt(Pmax))
    f = np.zeros((k))
    w = np.zeros((k))
    for i in range(k):
        f[i] = H[i,i] * b[i] /(np.square(H[i,:]) @ np.square(b) + var_noise)
        w[i] = 1 / (1 - f[i] * b[i] * H[i, i])
        vnew = vnew + log2(w[i])

    VV = np.zeros(100)
    for iter in range(100):
        vold = vnew
        for i in range(k):
            btmp = w[i] * f[i] * H[i, i] / sum(w * np.square(f) * np.square(H[:, i]))
            b[i] = min(btmp, np.sqrt(Pmax)) + max(btmp, 0) - btmp

        vnew = 0
        for i in range(k):
            f[i] = H[i, i] * b[i] / ((np.square(H[i, :])) @ (np.square(b)) + var_noise)
            w[i] = 1 / (1 - f[i] * b[i] * H[i, i])
            vnew = vnew + log2(w[i])

        VV[iter] = vnew
        if vnew - vold <= 1e-3:
            break

    p_opt = np.square(b)
    return p_opt

######################## algorithm end ####################################
###############################  For creating excel sheets of value ####################################
wb1=load_workbook('receiver1.xlsx')          
ws1 = wb1.active
wb2=load_workbook('receiver2.xlsx')
ws2 = wb2.active
wb3=load_workbook('receiver3.xlsx')
ws3 = wb3.active
wb4=load_workbook('receiver4.xlsx')
ws4 = wb4.active

for iteration in range(100) :
  rx_loc, tx_loc = loc_init(d2d_dist, num_d2d)           # initializing tx-rx pairs location 
  cap = np.zeros(num_d2d)
  a=0
  b=0
  c=0
  d=0
  for sim_time  in range(1000) :
     h=np.zeros((num_d2d,num_d2d))
     for i in range(num_d2d) :
      for j in range(num_d2d) :
        x = 0.5 * np.random.randn(num_d2d, num_d2d) ** 2 + 0.5 * np.random.randn(num_d2d, num_d2d) ** 2
        h[i,j]=np.sum(x)
     for k in range(num_d2d):
       s = 1
       for l in range(num_d2d):
         if l!=k :
          s = s+h[k,l]**2*p_opt[l]

     rx_loc, tx_loc = get_loc(rx_loc, tx_loc,num_d2d)
     p_opt=WMMSE(num_d2d,h,p_t,1)
     for i in range(num_d2d) :
      for j in range(num_d2d) :
        d =  np.linalg.norm(tx_loc[i] - rx_loc[j])  
        if d < 100 :
           if j==0:
            cap[j]+=log2(1+h[i,j]**2*p_opt[j]/s)
            a+=1
           elif j==1:
             cap[j] += log2(1+h[i,j]**2*p_opt[j]/s)
             b+=1
           elif j==2:
            cap[j] += log2(1+h[i,j]**2*p_opt[j]/s)
            c+=1
           elif j==3:
            cap[j]+=log2(1+h[i,j]**2*p_opt[j]/s)
            d+=1 
          
  for m in range(num_d2d) :
    if m==0 :
      cap[m]=cap[m]/a
      ws1.append([cap[m]])
    elif m==1:
      cap[m]=cap[m]/b
      ws2.append([cap[m]])
    elif m==2:
       cap[m]=cap[m]/c
       ws3.append([cap[m]])
    elif m==3:
      cap[m]=cap[m]/d
      ws4.append([cap[m]]) 

wb1.save('receiver1.xlsx')
wb2.save('receiver2.xlsx')
wb3.save('receiver3.xlsx')
wb4.save('receiver4.xlsx')
