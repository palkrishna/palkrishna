import pandas as pd
from openpyxl import workbook,load_workbook
import channel_initialization as ch
import numpy as np
####################### simulation parameters ########################
bw = 1*10**9              # Bandwidth = 1GHz
p_t_dB = 23               # maximum d2d transmit power in dB
p_t = 10**(p_t_dB/10)      
C_frequency = 28e9         # carrier frequency = 28 GHz
num_d2d =4
Noise = -174               # noise = -174 dBm
s= 10**(-174/10)

#########################  Assigning equal power of 
d2d_power =np.ones(num_d2d)*(23/4)
sinr = np.zeros(num_d2d)
cap = np.zeros(num_d2d)
###############################  For creating excel sheets of value ####################################
wb1=load_workbook('receiver1.xlsx')          
ws1 = wb1.active
wb2=load_workbook('receiver2.xlsx')
ws2 = wb2.active
wb3=load_workbook('receiver3.xlsx')
ws3 = wb3.active
wb4=load_workbook('receiver4.xlsx')
ws4 = wb4.active


for z in range(51) :
  h=ch.ch_gen(20,num_d2d,C_frequency)
  for i in range(num_d2d) :

    x = h[i,i]*d2d_power[i]/s
    sinr[i] = 10*log10(x)
    cap[i] = bw*log2(1+x)
    if i==0 :
      ws1.append([cap[i]])
    elif i==1:
       ws2.append([cap[i]])
    elif i==2:
       ws3.append([cap[i]])
    elif i==3:
      ws4.append([cap[i]])

wb1.save('receiver1.xlsx')
wb2.save('receiver2.xlsx')
wb3.save('receiver3.xlsx')
wb4.save('receiver4.xlsx')